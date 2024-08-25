import xml.dom.minidom
from openpyxl import Workbook
import os
import html
from bs4 import BeautifulSoup


def extrair_anos_inicio(doc):
    atuacoes_profissionais = doc.getElementsByTagName("ATUACAO-PROFISSIONAL")

    for atuacao in atuacoes_profissionais:
        atuacaoN = atuacao.getAttribute("SEQUENCIA-IMPORTANCIA")
        if atuacaoN == "1":
            vinculos = atuacao.getElementsByTagName("VINCULOS")
            if vinculos:
                vinculo = vinculos[0]
                ano_inicio = vinculo.getAttribute("ANO-INICIO")
                return ano_inicio
            else:
                return "não preenchido"


def dados_gerais(doc,  id, tipo):
    ano_inicio = extrair_anos_inicio(doc)
    elemento = doc.getElementsByTagName("CURRICULO-VITAE")[0]
    id_lattes = elemento.getAttribute("NUMERO-IDENTIFICADOR")
    elemento = doc.getElementsByTagName("DADOS-GERAIS")[0]
    nome = elemento.getAttribute("NOME-COMPLETO")
    print(nome)
    if (tipo == 1):
        
        return id, nome
    estado = elemento.getAttribute("UF-NASCIMENTO")
    cdd = elemento.getAttribute("CIDADE-NASCIMENTO")
    elemento = elemento.getElementsByTagName("ENDERECO")[0]
    elemento = elemento.getElementsByTagName("ENDERECO-PROFISSIONAL")[0]
    nome_inst = elemento.getAttribute("NOME-INSTITUICAO-EMPRESA")
    nome_orgao = elemento.getAttribute("NOME-ORGAO")
    tel = elemento.getAttribute("TELEFONE")
    cdd_inst = elemento.getAttribute("CIDADE")

    return id, id_lattes, nome, estado, cdd, nome_inst, ano_inicio, nome_orgao, cdd_inst, tel


def dados_formacao(tipo, doc, id):
    formacao = doc.getElementsByTagName("FORMACAO-ACADEMICA-TITULACAO")[0]
    titulo = ""
    orientador = ""
    instituicao = ""
    curso = ""

    if tipo == 0:
        elementos = formacao.getElementsByTagName("GRADUACAO")
    elif tipo == 1:  # Mestrado
        elementos = formacao.getElementsByTagName("MESTRADO")
    elif tipo == 2:  # Doutorado
        elementos = formacao.getElementsByTagName("DOUTORADO")
    elif tipo == 3:
        elementos = formacao.getElementsByTagName("ESPECIALIZACAO")
    elif tipo == 4:
        elementos = formacao.getElementsByTagName("POS-DOUTORADO")

    for elemento in elementos:
        if tipo == 0:
            titulo = elemento.getAttribute(
                "TITULO-DO-TRABALHO-DE-CONCLUSAO-DE-CURSO")
            orientador = elemento.getAttribute("NOME-DO-ORIENTADOR")
            instituicao = elemento.getAttribute("NOME-INSTITUICAO")
            curso = elemento.getAttribute("NOME-CURSO")
            return id, titulo, orientador, instituicao, curso

        elif tipo == 1 or tipo == 2:  # Mestrado ou Doutorado
            titulo = elemento.getAttribute("TITULO-DA-DISSERTACAO-TESE")
            orientador = elemento.getAttribute("NOME-COMPLETO-DO-ORIENTADOR")
            instituicao = elemento.getAttribute("NOME-INSTITUICAO")
            curso = elemento.getAttribute("NOME-CURSO")
            areas_conhecimento = elemento.getElementsByTagName(
                "AREA-DO-CONHECIMENTO-1")
            if areas_conhecimento:
                area_conhecimento = areas_conhecimento[0].getAttribute(
                    "NOME-GRANDE-AREA-DO-CONHECIMENTO")
                sub_area_conhecimento = areas_conhecimento[0].getAttribute(
                    "NOME-DA-SUB-AREA-DO-CONHECIMENTO")
            else:
                area_conhecimento = ""
                sub_area_conhecimento = ""
            ano_inicio = elemento.getAttribute("ANO-DE-INICIO")
            ano_conlusao = elemento.getAttribute("ANO-DE-CONCLUSAO")

            return id, titulo, orientador, instituicao, curso, area_conhecimento, sub_area_conhecimento, ano_inicio, ano_conlusao

        elif tipo == 3:
            titulo = elemento.getAttribute("TITULO-DA-MONOGRAFIA")
            orientador = elemento.getAttribute("NOME-DO-ORIENTADOR")
            instituicao = elemento.getAttribute("NOME-INSTITUICAO")
            curso = elemento.getAttribute("NOME-CURSO")
            return id, titulo, orientador, instituicao, curso

        elif tipo == 4:
            instituicao = elemento.getAttribute("NOME-INSTITUICAO")
            ano_conlusao = elemento.getAttribute("ANO-DE-CONCLUSAO")
            nome_agencia = elemento.getAttribute("NOME-AGENCIA")
            return id, instituicao, ano_conlusao, nome_agencia


def extrair_areas(wb, doc, id, linha):
    Gareas = []
    Area = []
    Sareas = []
    especialidade = []

    areas_de_atuacao = doc.getElementsByTagName("AREAS-DE-ATUACAO")
    if areas_de_atuacao:
        areas_de_atuacao = areas_de_atuacao[0]
    else:
        wb.append([0])
        return

    areas = areas_de_atuacao.getElementsByTagName("AREA-DE-ATUACAO")
    numero_de_areas = len(areas)

    wb.append([id, numero_de_areas])
    for area in areas:
        gNomeArea = area.getAttribute("NOME-GRANDE-AREA-DO-CONHECIMENTO")
        Gareas.append(gNomeArea)
        nomeArea = area.getAttribute("NOME-DA-AREA-DO-CONHECIMENTO")
        Area.append(nomeArea)
        subArea = area.getAttribute("NOME-DA-SUB-AREA-DO-CONHECIMENTO")
        Sareas.append(subArea)
        nomeEsp = area.getAttribute("NOME-DA-ESPECIALIDADE")
        especialidade.append(nomeEsp)

    for i in range(len(Gareas)):
        wb.cell(row=linha, column=3*i + 2, value=Gareas[i])
        wb.cell(row=linha, column=3*i + 3, value=Area[i])
        wb.cell(row=linha, column=3*i + 4, value=Sareas[i])
        wb.cell(row=linha, column=3*i + 5, value=especialidade[i])


def extrair_premios(doc, id):
    premios_titulos = doc.getElementsByTagName("PREMIOS-TITULOS")
    if premios_titulos:
        premios_titulos = premios_titulos[0]
    else:
        return

    premios = premios_titulos.getElementsByTagName("PREMIO-TITULO")

    nome1, nome_entidade1, ano1 = "", "", ""
    nome2, nome_entidade2, ano2 = "", "", ""
    nome3, nome_entidade3, ano3 = "", "", ""

    i = 1
    for premio in premios:
        if i == 1:
            nome1 = premio.getAttribute("NOME-DO-PREMIO-OU-TITULO")
            nome_entidade1 = premio.getAttribute("NOME-DA-ENTIDADE-PROMOTORA")
            ano1 = premio.getAttribute("ANO-DA-PREMIACAO")
        elif i == 2:
            nome2 = premio.getAttribute("NOME-DO-PREMIO-OU-TITULO")
            nome_entidade2 = premio.getAttribute("NOME-DA-ENTIDADE-PROMOTORA")
            ano2 = premio.getAttribute("ANO-DA-PREMIACAO")
        elif i == 3:
            nome3 = premio.getAttribute("NOME-DO-PREMIO-OU-TITULO")
            nome_entidade3 = premio.getAttribute("NOME-DA-ENTIDADE-PROMOTORA")
            ano3 = premio.getAttribute("ANO-DA-PREMIACAO")
        i += 1

        if i > 3:
            break

    return id, nome1, nome_entidade1, ano1, nome2, nome_entidade2, ano2, nome3, nome_entidade3, ano3


def extrair_artigo(wb, doc, id, linha):
    aux = linha
    artigos_publicados = doc.getElementsByTagName("ARTIGOS-PUBLICADOS")
    anos = []
    periodicos = []
    if artigos_publicados:
        artigos_publicados = artigos_publicados[0]
    else:
        wb.append([0])
        return

    artigos = artigos_publicados.getElementsByTagName("ARTIGO-PUBLICADO")

    numero_de_artigos = len(artigos)

    wb.append([id, numero_de_artigos])
    for artigo in artigos:
        dados = artigo.getElementsByTagName("DADOS-BASICOS-DO-ARTIGO")
        for dado in dados:
            ano = dado.getAttribute("ANO-DO-ARTIGO")
            anos.append(ano)
        detalhes = artigo.getElementsByTagName("DETALHAMENTO-DO-ARTIGO")
        for detalhe in detalhes:
            periodico = detalhe.getAttribute("TITULO-DO-PERIODICO-OU-REVISTA")
            periodicos.append(periodico)
    k = 1
    j = 0

    for index, valor in enumerate(anos, start=1):
        if index == 1:
            index += 1
        else:
            index = index + k
        wb.cell(row=aux, column=index + 1, value=valor)

        wb.cell(row=aux, column=index + 2, value=periodicos[j])
        k += 1
        j += 1


def decode_entities(text):

    return html.unescape(text)


def clean_html(raw_html):

    soup = BeautifulSoup(raw_html, "html.parser")
    return soup.get_text()


def extrair_artigo_tiulo(wb, doc, id, linha):

    aux = linha
    titles = []
    try:
        # Extract the full name
        general_data = doc.getElementsByTagName("DADOS-GERAIS")[0]
        full_name = general_data.getAttribute("NOME-COMPLETO")
    except (IndexError, AttributeError):
        print(
            f"Error: 'DADOS-GERAIS' tag or 'NOME-COMPLETO' attribute missing in document ID {id}")
        return

    print(id)
    wb.append([id, full_name])

    try:
        published_articles = doc.getElementsByTagName("ARTIGOS-PUBLICADOS")[0]
        articles = published_articles.getElementsByTagName("ARTIGO-PUBLICADO")
    except IndexError:
        wb.append([0])
        return

    for article in articles:
        basic_data = article.getElementsByTagName("DADOS-BASICOS-DO-ARTIGO")
        for data in basic_data:
            article_title = decode_entities(clean_html(
                data.getAttribute("TITULO-DO-ARTIGO")))
            titles.append(article_title)

    for index, title in enumerate(titles, start=2):
        wb.cell(row=aux, column=index + 1, value=title)


def extrair_trabalhos(wb, doc, id, tipo, linha):
    producao_bibliografica = doc.getElementsByTagName("PRODUCAO-BIBLIOGRAFICA")
    if producao_bibliografica:
        producao_bibliografica = producao_bibliografica[0]
        trabalho_publicados = producao_bibliografica.getElementsByTagName(
            "TRABALHOS-EM-EVENTOS")
    else:
        wb.append([id, 0])
        return

    anos = []
    eventos = []
    if trabalho_publicados:
        trabalho_publicados = trabalho_publicados[0]
    else:
        wb.append([id, 0])
        return

    trabalhos = trabalho_publicados.getElementsByTagName("TRABALHO-EM-EVENTOS")
    numero_de_trabalhos = 0

    for trabalho in trabalhos:
        dados = trabalho.getElementsByTagName("DADOS-BASICOS-DO-TRABALHO")
        if tipo == 1:
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "COMPLETO":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO-DO-TRABALHO")
                    anos.append(ano)
            if natureza == "COMPLETO":
                detalhes = trabalho.getElementsByTagName(
                    "DETALHAMENTO-DO-TRABALHO")
                for detalhe in detalhes:
                    evento = detalhe.getAttribute("NOME-DO-EVENTO")
                    eventos.append(evento)
        elif tipo == 2:
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "RESUMO_EXPANDIDO":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO-DO-TRABALHO")
                    anos.append(ano)
            if natureza == "RESUMO_EXPANDIDO":
                detalhes = trabalho.getElementsByTagName(
                    "DETALHAMENTO-DO-TRABALHO")
                for detalhe in detalhes:
                    evento = detalhe.getAttribute("NOME-DO-EVENTO")
                    eventos.append(evento)
        elif tipo == 3:
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "RESUMO":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO-DO-TRABALHO")
                    anos.append(ano)
            if natureza == "RESUMO":
                detalhes = trabalho.getElementsByTagName(
                    "DETALHAMENTO-DO-TRABALHO")
                for detalhe in detalhes:
                    evento = detalhe.getAttribute("NOME-DO-EVENTO")
                    eventos.append(evento)

    wb.append([id, numero_de_trabalhos])

    k = 1
    j = 0
    for index, valor in enumerate(anos, start=1):
        if index == 1:
            index += 1
        else:
            index = index + k
        wb.cell(row=linha, column=index + 1, value=valor)
        wb.cell(row=linha, column=index + 2, value=eventos[j])
        k += 1
        j += 1


def Producao_tecnica(wb, doc, id, tipo, linha):
    anos = []
    eventos = []
    naturezas = []
    numero_de_trabalhos = 0
    producao_tecnica = doc.getElementsByTagName("PRODUCAO-TECNICA")
    if producao_tecnica:
        doc = producao_tecnica[0]
    else:
        wb.append([id, 0])
        return

    if tipo == 1:
        demais_tipos = doc.getElementsByTagName(
            "DEMAIS-TIPOS-DE-PRODUCAO-TECNICA")
        if demais_tipos:
            doc = demais_tipos[0]
        else:
            wb.append([id, 0])
            return
        trabalhos = doc.getElementsByTagName("APRESENTACAO-DE-TRABALHO")

        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DA-APRESENTACAO-DE-TRABALHO")
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                numero_de_trabalhos += 1
                naturezas.append(natureza)
                ano = dado.getAttribute("ANO")
                anos.append(ano)
                detalhes = trabalho.getElementsByTagName(
                    "DETALHAMENTO-DA-APRESENTACAO-DE-TRABALHO")
                for detalhe in detalhes:
                    evento = detalhe.getAttribute("NOME-DO-EVENTO")
                    eventos.append(evento)
    elif tipo == 2:
        processos_ou_tecnicas = doc.getElementsByTagName(
            "PROCESSOS-OU-TECNICAS")
        if not processos_ou_tecnicas:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in processos_ou_tecnicas:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DO-PROCESSOS-OU-TECNICAS")
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                numero_de_trabalhos += 1
                naturezas.append(natureza)
                ano = dado.getAttribute("ANO")
                anos.append(ano)
                evento = dado.getAttribute("TITULO-DO-PROCESSO")
                eventos.append(evento)

    wb.append([id, numero_de_trabalhos])

    k = 1
    j = 0
    for index, valor in enumerate(anos, start=1):
        if index == 1:
            index += 1
        else:
            index = index + k
        wb.cell(row=linha, column=index + 1, value=naturezas[j])
        wb.cell(row=linha, column=index + 2, value=valor)
        wb.cell(row=linha, column=index + 3, value=eventos[j])
        k += 2
        j += 1


def Bancas(wb, doc, id, tipo, linha):
    anos = []
    tipos = []
    numero_de_trabalhos = 0
    doc = doc.getElementsByTagName("DADOS-COMPLEMENTARES")
    if doc:
        doc = doc[0]

    doc = doc.getElementsByTagName("PARTICIPACAO-EM-BANCA-TRABALHOS-CONCLUSAO")
    if doc:
        doc = doc[0]
    else:
        wb.append([id, 0])
        return

    if tipo == 1:
        trabalhos = doc.getElementsByTagName(
            "PARTICIPACAO-EM-BANCA-DE-MESTRADO")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-MESTRADO")
            for dado in dados:
                type = dado.getAttribute("TIPO")
                numero_de_trabalhos += 1
                tipos.append(type)
                ano = dado.getAttribute("ANO")
                anos.append(ano)

    if tipo == 2:
        trabalhos = doc.getElementsByTagName(
            "PARTICIPACAO-EM-BANCA-DE-DOUTORADO")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DA-PARTICIPACAO-EM-BANCA-DE-DOUTORADO")
            for dado in dados:
                type = dado.getAttribute("TITULO")
                numero_de_trabalhos += 1
                tipos.append(type)
                ano = dado.getAttribute("ANO")
                anos.append(ano)

    k = 1
    j = 0
    wb.append([id, numero_de_trabalhos])
    for index, valor in enumerate(anos, start=1):
        if index == 1:
            index += 1
        else:
            index = index + k
        if j < len(anos):
            wb.cell(row=linha, column=index + 1, value=valor)
        if j < len(tipos):
            wb.cell(row=linha, column=index + 2, value=tipos[j])
        k += 1
        j += 1


def orientacao(wb, doc, id, tipo, linha):
    anos = []
    tipos = []
    numero_de_trabalhos = 0
    outra_producao = doc.getElementsByTagName("OUTRA-PRODUCAO")
    if outra_producao:
        outra_producao = outra_producao[0]
    else:
        wb.append([id, 0])
        return

    orientacoes_concluidas = outra_producao.getElementsByTagName(
        "ORIENTACOES-CONCLUIDAS")
    if orientacoes_concluidas:
        orientacoes_concluidas = orientacoes_concluidas[0]
    else:
        wb.append([id, 0])
        return

    if tipo == 1:
        trabalhos = orientacoes_concluidas.getElementsByTagName(
            "ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
            for dado in dados:
                numero_de_trabalhos += 1
                ano = dado.getAttribute("ANO")
                anos.append(ano)
            detalhes = trabalho.getElementsByTagName(
                "DETALHAMENTO-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
            for detalhe in detalhes:
                tipo_orientacao = detalhe.getAttribute("TIPO-DE-ORIENTACAO")
                tipos.append(tipo_orientacao)

    elif tipo == 2:
        trabalhos = orientacoes_concluidas.getElementsByTagName(
            "ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO")
            for dado in dados:
                numero_de_trabalhos += 1
                ano = dado.getAttribute("ANO")
                anos.append(ano)
            detalhes = trabalho.getElementsByTagName(
                "DETALHAMENTO-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO")
            for detalhe in detalhes:
                tipo_orientacao = detalhe.getAttribute("TIPO-DE-ORIENTACAO")
                tipos.append(tipo_orientacao)

    elif tipo == 3:
        trabalhos = orientacoes_concluidas.getElementsByTagName(
            "OUTRAS-ORIENTACOES-CONCLUIDAS")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            detalhes = trabalho.getElementsByTagName(
                "DETALHAMENTO-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "INICIACAO_CIENTIFICA":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO")
                    anos.append(ano)
                    for detalhe in detalhes:
                        tipo_orientacao = detalhe.getAttribute(
                            "TIPO-DE-ORIENTACAO-CONCLUIDA")
                        tipos.append(tipo_orientacao)

    elif tipo == 4:
        trabalhos = orientacoes_concluidas.getElementsByTagName(
            "OUTRAS-ORIENTACOES-CONCLUIDAS")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            detalhes = trabalho.getElementsByTagName(
                "DETALHAMENTO-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "ORIENTACAO-DE-OUTRA-NATUREZA":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO")
                    anos.append(ano)
                    for detalhe in detalhes:
                        tipo_orientacao = detalhe.getAttribute(
                            "TIPO-DE-ORIENTACAO-CONCLUIDA")
                        tipos.append(tipo_orientacao)

    elif tipo == 5:
        trabalhos = orientacoes_concluidas.getElementsByTagName(
            "OUTRAS-ORIENTACOES-CONCLUIDAS")
        if not trabalhos:
            wb.append([id, numero_de_trabalhos])
            return
        for trabalho in trabalhos:
            dados = trabalho.getElementsByTagName(
                "DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            detalhes = trabalho.getElementsByTagName(
                "DETALHAMENTO-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
            for dado in dados:
                natureza = dado.getAttribute("NATUREZA")
                if natureza == "TRABALHO_DE_CONCLUSAO_DE_CURSO_GRADUACAO":
                    numero_de_trabalhos += 1
                    ano = dado.getAttribute("ANO")
                    anos.append(ano)
                    for detalhe in detalhes:
                        tipo_orientacao = detalhe.getAttribute(
                            "TIPO-DE-ORIENTACAO-CONCLUIDA")
                        tipos.append(tipo_orientacao)

    wb.append([id, numero_de_trabalhos])
    k = 1
    j = 0
    for index, valor in enumerate(anos, start=1):
        if index == 1:
            index += 1
        else:
            index = index + k
        if j < len(anos):
            wb.cell(row=linha, column=index + 1, value=valor)
        if j < len(tipos):
            wb.cell(row=linha, column=index + 2, value=tipos[j])
        k += 1
        j += 1


def extrair_projetos(wb, doc, id, tipo, linha):
    atuacao = doc.getElementsByTagName("ATUACAO-PROFISSIONAL")
    aux = 0
    cont = 0
    anos_proj = []

    for att in atuacao:
        att_projetos = att.getElementsByTagName(
            "ATIVIDADES-DE-PARTICIPACAO-EM-PROJETO")
        if not att_projetos:
            aux += 1
            continue

        participacoes = att_projetos[0].getElementsByTagName(
            "PARTICIPACAO-EM-PROJETO")

        if not participacoes:
            aux += 1
            continue

        for participacao in participacoes:
            projetos_pesquisa = participacao.getElementsByTagName(
                "PROJETO-DE-PESQUISA")
            for projeto in projetos_pesquisa:
                natureza = projeto.getAttribute("NATUREZA")
                if tipo == 1:
                    if natureza == "PESQUISA":
                        ano_inicio = projeto.getAttribute("ANO-INICIO")
                        ano_fim = projeto.getAttribute("ANO-FIM")
                        if not ano_fim:
                            ano_fim = "atual"
                        ano_inicio_fim = ano_inicio + "-" + ano_fim
                        anos_proj.append(ano_inicio_fim)
                        cont += 1
                elif tipo == 2:
                    if natureza == "DESENVOLVIMENTO":
                        ano_inicio = projeto.getAttribute("ANO-INICIO")
                        ano_fim = projeto.getAttribute("ANO-FIM")
                        if not ano_fim:
                            ano_fim = "atual"
                        ano_inicio_fim = ano_inicio + "-" + ano_fim
                        anos_proj.append(ano_inicio_fim)
                        cont += 1
                elif tipo == 3:
                    if natureza == "EXTENSAO":
                        ano_inicio = projeto.getAttribute("ANO-INICIO")
                        ano_fim = projeto.getAttribute("ANO-FIM")
                        if not ano_fim:
                            ano_fim = "atual"
                        ano_inicio_fim = ano_inicio + "-" + ano_fim
                        anos_proj.append(ano_inicio_fim)
                        cont += 1
                elif tipo == 4:
                    if natureza == "ENSINO":
                        ano_inicio = projeto.getAttribute("ANO-INICIO")
                        ano_fim = projeto.getAttribute("ANO-FIM")
                        if not ano_fim:
                            ano_fim = "atual"
                        ano_inicio_fim = ano_inicio + "-" + ano_fim
                        anos_proj.append(ano_inicio_fim)
                        cont += 1
    if aux != len(atuacao):
        wb.append([id, cont])
        for index, valor in enumerate(anos_proj, start=1):
            index = index + 2
            wb.cell(row=linha, column=index, value=valor)
    else:
        wb.append([id, 0])


def extrair_quantativamente(doc, id):
    artigos_publicados = doc.getElementsByTagName("ARTIGOS-PUBLICADOS")
    numero_de_artigos = 0 if not artigos_publicados else len(
        artigos_publicados[0].getElementsByTagName("ARTIGO-PUBLICADO"))

    numero_de_trabalhos_completos = 0
    numero_de_trabalhos_expandido = 0
    numero_de_trabalhos = 0

    tb = doc.getElementsByTagName("PRODUCAO-BIBLIOGRAFICA")
    if tb:
        trabalho_publicados = tb[0].getElementsByTagName(
            "TRABALHOS-EM-EVENTOS")
        if trabalho_publicados:
            trabalhos = trabalho_publicados[0].getElementsByTagName(
                "TRABALHO-EM-EVENTOS")
            for i in range(3):
                for tipo in trabalhos:
                    dados = tipo.getElementsByTagName(
                        "DADOS-BASICOS-DO-TRABALHO")
                    for dado in dados:
                        natureza = dado.getAttribute("NATUREZA")
                        if i == 0 and natureza == "COMPLETO":
                            numero_de_trabalhos_completos += 1
                        elif i == 1 and natureza == "RESUMO_EXPANDIDO":
                            numero_de_trabalhos_expandido += 1
                        elif i == 2 and natureza == "RESUMO":
                            numero_de_trabalhos += 1

    bancas_m = 0
    bancas_d = 0
    b = doc.getElementsByTagName("DADOS-COMPLEMENTARES")
    if b:
        b = b[0].getElementsByTagName(
            "PARTICIPACAO-EM-BANCA-TRABALHOS-CONCLUSAO")
        if b:
            bancasM = doc.getElementsByTagName(
                "PARTICIPACAO-EM-BANCA-DE-MESTRADO")
            bancasD = doc.getElementsByTagName(
                "PARTICIPACAO-EM-BANCA-DE-DOUTORADO")
            bancas_m = len(bancasM) if bancasM else 0
            bancas_d = len(bancasD) if bancasD else 0

    ori_M = 0
    ori_D = 0
    ori_IC = 0
    ori_G = 0
    o = doc.getElementsByTagName("OUTRA-PRODUCAO")
    if o:
        Mes = doc.getElementsByTagName("ORIENTACOES-CONCLUIDAS-PARA-MESTRADO")
        Dou = doc.getElementsByTagName("ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO")
        ori_M = len(Mes) if Mes else 0
        ori_D = len(Dou) if Dou else 0

        for i in range(2):
            trabalhos = doc.getElementsByTagName(
                "OUTRAS-ORIENTACOES-CONCLUIDAS")
            if trabalhos:
                for trabalho in trabalhos:
                    dados = trabalho.getElementsByTagName(
                        "DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS")
                    for dado in dados:
                        natureza = dado.getAttribute("NATUREZA")
                        if i == 1 and natureza == "INICIACAO_CIENTIFICA":
                            ori_IC += 1
                        elif i == 0 and natureza == "TRABALHO_DE_CONCLUSAO_DE_CURSO_GRADUACAO":
                            ori_G += 1

    numero_de_pesquisa = 0
    numero_de_desenvolvimento = 0
    numero_de_extensao = 0
    numero_de_ensino = 0

    atuacao = doc.getElementsByTagName("ATUACAO-PROFISSIONAL")
    for att in atuacao:
        att_projetos = att.getElementsByTagName(
            "ATIVIDADES-DE-PARTICIPACAO-EM-PROJETO")
        if att_projetos:
            participacoes = att_projetos[0].getElementsByTagName(
                "PARTICIPACAO-EM-PROJETO")
            if participacoes:
                for participacao in participacoes:
                    projetos_pesquisa = participacao.getElementsByTagName(
                        "PROJETO-DE-PESQUISA")
                    for projeto in projetos_pesquisa:
                        natureza = projeto.getAttribute("NATUREZA")
                        if natureza == "PESQUISA":
                            numero_de_pesquisa += 1
                        elif natureza == "DESENVOLVIMENTO":
                            numero_de_desenvolvimento += 1
                        elif natureza == "EXTENSAO":
                            numero_de_extensao += 1
                        elif natureza == "ENSINO":
                            numero_de_ensino += 1

    return id, numero_de_pesquisa, numero_de_extensao, numero_de_ensino, numero_de_desenvolvimento, numero_de_artigos, numero_de_trabalhos_expandido, numero_de_trabalhos_completos, numero_de_trabalhos, bancas_m, bancas_d, ori_IC, ori_G, ori_M, ori_D


def criar_planilhas(wb):
    wb.create_sheet('Graduação')
    wb.create_sheet('Especialização')
    wb.create_sheet('Mestrado')
    wb.create_sheet('Doutorado')
    wb.create_sheet('Pós-Doutorado')
    wb.create_sheet('Areas de Atuação')
    wb.create_sheet('Planilha quantitativa')
    wb.create_sheet('Projetos de pesquisa')
    wb.create_sheet('Projetos de desenvolvimento')
    wb.create_sheet('Projetos de extensão')
    wb.create_sheet('Projetos de ensino')
    wb.create_sheet('Artigos em periódicos')
    wb.create_sheet('Trabalhos completos em Eventos')
    wb.create_sheet('Resumos Expandidos')
    wb.create_sheet('Resumos Publicados')
    wb.create_sheet('Apresentação de trabalho')
    wb.create_sheet('Processos ou tecnicas')
    wb.create_sheet('Part. Bancas Mestrado')
    wb.create_sheet('Part. Bancas Doutorado')
    wb.create_sheet('Orientações Graduação')
    wb.create_sheet('Orientações Mestrado')
    wb.create_sheet('Orientações Doutorado')
    wb.create_sheet('Orientações de I.C.')
    wb.create_sheet('Outras Orientações')
    wb.create_sheet('Premios e Titulos')
    wb.create_sheet('Dados Quimica')

    planilha = wb['Dados Gerais']
    planilha.append(["id", "ID Lattes", "Nome", "UF-Nascimento", "Cidade Nascimento",
                    "Instituição Atual", "Ano de ingresso", "Nome do Orgão", "Cidade da instituição", "Telefone Orgão"])
    planilha = wb['Graduação']
    planilha.append(["id", "Título do trabalho",
                    "Orientador", "Instituição", "Curso"])
    planilha = wb['Especialização']
    planilha.append(["id", "Título da Monografia",
                    "Orientador", "Instituição", "Curso"])
    planilha = wb['Mestrado']
    planilha.append(["id", "Título da dissertação", "Orientador", "Instituição",
                    "Curso", "Grande Area do conhecimento", "Sub area do conhecimento", "ano de Incio", "Ano de Conclusão"])
    planilha = wb['Doutorado']
    planilha.append(["id", "Título da dissertação", "Orientador", "Instituição",
                    "Curso", "Grande Area do conhecimento", "Sub area do conhecimento", "ano de Incio", "Ano de Conclusão"])
    planilha = wb['Pós-Doutorado']
    planilha.append(["id", "Nome instituição",
                    "Nome Agencia", "Ano de conclusão"])
    planilha = wb['Areas de Atuação']
    planilha.append(["id", "Area de Atuação: Nome da grande area", "Area de Atuação: Nome da area",
                    "Area de Atuação: Nome da sub area", "Area de Atuação: Nome da especialidade"])
    planilha = wb['Planilha quantitativa']
    planilha.append(["id", "P. Pesquisa", "P. Extensão", "P. Ensino", "P. Desenvolvimento", "Artigos", "Resumos Expandidos", "Trabalhos em Eventos",
                    "Resumos Publicados", "Bancas Mr.", "Bancas Dr.", "Orientações I.C.", "Orientações G.", "Orientações Mr.", "Orientações Dr."])
    planilha = wb['Projetos de pesquisa']
    planilha.append(["id", "N° de Projetos"])
    planilha = wb['Projetos de desenvolvimento']
    planilha.append(["id", "N° de Projetos"])
    planilha = wb['Projetos de extensão']
    planilha.append(["id", "N° de Projetos"])
    planilha = wb['Projetos de ensino']
    planilha.append(["id", "N° de projetos"])
    planilha = wb['Artigos em periódicos']
    planilha.append(["id", "N° de artigos"])
    planilha = wb['Trabalhos completos em Eventos']
    planilha.append(["id", "N° de Trabalhos completos"])
    planilha = wb['Resumos Expandidos']
    planilha.append(["id", "N° de resumos"])
    planilha = wb['Resumos Publicados']
    planilha.append(["id", "N° de resumos"])
    planilha = wb['Apresentação de trabalho']
    planilha.append(["id", "N° de apresentações"])
    planilha = wb['Processos ou tecnicas']
    planilha.append(["id", "N° de processos/tecnicas"])
    planilha = wb['Part. Bancas Mestrado']
    planilha.append(["id", "N° de Part."])
    planilha = wb['Part. Bancas Doutorado']
    planilha.append(["id", "N° de Part."])
    planilha = wb['Orientações Graduação']
    planilha.append(["id", "N° de Orientações"])
    planilha = wb['Orientações Mestrado']
    planilha.append(["id", "N° de Orientações"])
    planilha = wb['Orientações Doutorado']
    planilha.append(["id", "N° de Orientações"])
    planilha = wb['Orientações de I.C.']
    planilha.append(["id", "N° de Orientações"])
    planilha = wb['Outras Orientações']
    planilha.append(["id", "N° de Orientações"])
    planilha = wb['Premios e Titulos']
    planilha.append(["id", "Premio 1: nome do premio/titulo", "Premio 1: entidade promotora", "Premio 1: Ano da premiação", "Premio 2: nome do premio/titulo",
                    "Premio 2: entidade promotora", "Premio 2: Ano da premiação", "Premio 3: nome do premio/titulo", "Premio 3: entidade promotora", "Premio 3: Ano da premiação"])
    planilha = wb['Dados Quimica']
    planilha.append(["id", "Nome", "Titulo de Artigos"])


def add_dados(wb, tipo, id, dados=None):
    if tipo == 0:
        planilha = wb['Dados Gerais']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 1:
        planilha = wb['Graduação']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 3:
        planilha = wb['Doutorado']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 2:
        planilha = wb['Mestrado']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 4:
        planilha = wb['Especialização']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 5:
        planilha = wb['Areas de Atuação']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 6:
        planilha = wb['Premios e Titulos']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 7:
        planilha = wb['Artigos']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 8:
        planilha = wb['Pós-Doutorado']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 9:
        planilha = wb['Patentes']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 10:
        planilha = wb['Planilha quantitativa']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])
    elif tipo == 11:
        planilha = wb['Dados Quimica']
        if dados:
            planilha.append(list(dados))
        else:
            planilha.append([id])


def nova_planilha(wb):
    diretorio = "Curriculos_quimica"
    arquivos = os.listdir(diretorio)
    id = 0
    linha = 1

    for arquivo in arquivos:
        id = id + 1
        linha += 1

        caminho_arquivo = os.path.join(diretorio, arquivo)
        with open(caminho_arquivo, 'r') as arquivo_xml:
            conteudo_xml = arquivo_xml.read()
        doc = xml.dom.minidom.parseString(conteudo_xml)
        extrair_artigo_tiulo(wb["Dados Quimica"], doc, id, linha)


def main():
    wb = Workbook()
    wb.active.title = 'Dados Gerais'
    id = 0
    linha = 1
    criar_planilhas(wb)

    diretorio = "Curriculos_quimica"
    arquivos = os.listdir(diretorio)

    for arquivo in arquivos:
        id = id + 1
        caminho_arquivo = os.path.join(diretorio, arquivo)
        with open(caminho_arquivo, 'r') as arquivo_xml:
            conteudo_xml = arquivo_xml.read()
        doc = xml.dom.minidom.parseString(conteudo_xml)
        add_dados(wb, 0, id, dados_gerais(doc, id, 0))
        add_dados(wb, 1, id, dados_formacao(0, doc, id))
        add_dados(wb, 3, id, dados_formacao(2, doc, id))
        add_dados(wb, 2, id, dados_formacao(1, doc, id))
        add_dados(wb, 4, id, dados_formacao(3, doc, id))
        add_dados(wb, 6, id, extrair_premios(doc, id))
        add_dados(wb, 8, id, dados_formacao(4, doc, id))
        # add_dados(wb, 9, id, extrair_patentes(doc, id))
        linha = linha + 1
        extrair_projetos(wb['Projetos de pesquisa'], doc, id, 1, linha)
        extrair_projetos(wb['Projetos de desenvolvimento'], doc, id, 2, linha)
        extrair_projetos(wb['Projetos de extensão'], doc, id, 3, linha)
        extrair_projetos(wb['Projetos de ensino'], doc, id, 4, linha)
        extrair_artigo(wb['Artigos em periódicos'], doc, id, linha)
        extrair_trabalhos(
            wb['Trabalhos completos em Eventos'], doc, id, 1, linha)
        extrair_trabalhos(wb['Resumos Expandidos'], doc, id, 2, linha)
        extrair_trabalhos(wb['Resumos Publicados'], doc, id, 3, linha)
        Producao_tecnica(wb['Apresentação de trabalho'], doc, id, 1,  linha)
        Producao_tecnica(wb['Processos ou tecnicas'], doc, id, 2,  linha)
        Bancas(wb['Part. Bancas Mestrado'], doc, id, 1, linha)
        Bancas(wb['Part. Bancas Doutorado'], doc, id, 2, linha)
        orientacao(wb['Orientações Mestrado'], doc, id, 1, linha)
        orientacao(wb['Orientações Doutorado'], doc, id, 2, linha)
        orientacao(wb['Orientações de I.C.'], doc, id, 3, linha)
        orientacao(wb['Outras Orientações'], doc, id, 4, linha)
        orientacao(wb['Orientações Graduação'], doc, id, 5, linha)
        extrair_areas(wb['Areas de Atuação'], doc, id, linha)
        add_dados(wb, 10, id, extrair_quantativamente(doc, id))
        print(linha)

    wb.save('Dados Quimica.xlsx')


if __name__ == '__main__':
    main()
